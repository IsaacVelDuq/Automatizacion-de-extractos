import pdfplumber as plumb
import re, os
import pandas as pd
from itertools import zip_longest
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl
from openpyxl.styles import Font

def normalize_table(table):
    headers = table[0]
    rows = table[1:]
    #cuando se lee la tabla desde pdfplumber, esta separa los datos por \n,
    # la siguiente función separa los datos

    split_rows = []
    for row in rows:
        split_cells = []
        for c in row:
            if c in [None, ""]:
                split_cells.append([""])
            else:
                texto = str(c)
                split_cells.append(texto.split("\n"))

    #con zip_longest se garantiza que todas las filas tengan la misma longitud y asi no romper el DF
        for values in zip_longest(*split_cells, fillvalue=""):
            split_rows.append(values)

    return pd.DataFrame(split_rows, columns=headers)


def adjust_cells(df):
    #Al leer las tablas con pdfplumber.read_tables(), algunos datos en las filas se desplazan hacia el siguiente campo, esta función normaliza los campos
    
    #como hay problemas en la lectura, esta condicion ayuda a que no entre al condicional 
    #las filas que contienen conversiona  moneda local en el campo descripción, 
    if ("Descripción" in df.columns) and ("to Fecha" in df.columns):

         #Ayuda a solucionar el problema de lectura en el desplazamiento de caracteres entre campos  
        def adjust_string(a,char):
            b=""
            if char in a:
                left,right = a.split(char,1)
                new_a = left
                new_b = right
            else:
                new_a = a
                new_b = b 
            return new_a,new_b
        #funciones que normalizan el desplazamiento de caracteres entre campos

        def process_row(row):
            desc = str(row["Descripción"]) or ""
            date = str(row["to Fecha"]) or ""
            if date.strip() != "" and desc and desc[0].isdigit():
                new_date= date + desc[0]
                new_desc = desc[1:]
            else:
                new_date=date
                new_desc=desc
            return pd.Series({"to Fecha": new_date, "Descripción":new_desc})
        df[["to Fecha", "Descripción"]] = df.apply(process_row, axis=1)

        def process_value(item):
            a= str(item["Valor a Pagar"])
            char = "$"
            left,right = adjust_string(a,char)
            new_valor= (str(item["Valor"]) + left).replace(char,"").strip()
            new_valor_pagar = right
            return pd.Series({"Valor":new_valor,"Valor a Pagar":new_valor_pagar})
        df[["Valor", "Valor a Pagar"]]= df.apply(process_value,axis=1)

        def process_balance(item):
            a=str(item["Saldo Pendiente"])
            char = "$"
            left,right = adjust_string(a,"$")
            new_valor=  str(item["Valor a Pagar"]) + left
            new_saldo = right
            return pd.Series({"Valor a Pagar":new_valor.replace(char,"").strip(), "Saldo Pendiente":new_saldo})
        df[["Valor a Pagar","Saldo Pendiente"]]= df.apply(process_balance,axis=1)

        #convierte una cadena con signo a la dereha en una cadena con signo a la izquierda
        def to_num(n):
            n=str(n).strip()
            if n.endswith(("+","-")):
                sign=n[-1]
                num = n[:-1]
                n= sign+num
            return n
        df["Valor a Pagar"]= df["Valor a Pagar"].apply(to_num)

    return df


def parse(df):
    #se renombran columnas ya que pdfplumber.read_tables() tiene problemas a leer los headers
    df.columns = [
        "Documen No.",
        "to Fecha",
        "Descripción",
        "Valor", 
        "Valor a Pagar",
        "Saldo Pendiente",
        "No.Cuota",
        "Cuota sPend.",
        "sTasa E.A.",
        "Tasa M.V"
    ]
    #pdf_read_tables() tiene fallos en la lectura de las primeras dos lineas

    df = df.drop([0,1]) 
    #se realiza un parseo a las columnas necesarias
    df["to Fecha"] = pd.to_datetime(df["to Fecha"], format="%Y%m%d").dt.date
    numeric_cols = ["Valor", "Valor a Pagar", "Saldo Pendiente", "No.Cuota", "Cuota sPend.", "sTasa E.A.", "Tasa M.V"]

    for col in numeric_cols: 
        df[col] = ( df[col].astype(str) .str.replace(r"[\",$]", "", regex=True).str.strip() )
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


    
def read_table(pdf_path):

    #Extrae tablas de un PDF y devuelve un DataFrame normalizado.
    pattern = re.compile(r"MOVIMIENTOS", re.IGNORECASE)
    dfs = []
    with plumb.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            if pattern.search(text):
                tables = page.extract_tables({
                    "vertical_strategy": "lines",      # usa las líneas verticales
                    "horizontal_strategy": "text",     # filas basadas en texto
                    "intersection_tolerance": 4,       # margen de error en intersecciones
                    "snap_tolerance": 3,               # ajusta bordes a líneas cercanas
                    "join_tolerance": 1,               # une fragmentos de texto dentro de la misma celda
                    "edge_min_length": 3,              # ignora líneas verticales muy cortas
                    "text_x_tolerance": 3              # controla separación horizontal del texto
                })
                
                for table in tables:
                    if len(table)>3 and len(table[0]) == 10: #si la tabla está vacía, el pdfplumber.extract_tables() lo lee con tres filas
                        df = normalize_table(table)
                        df = adjust_cells(df)                        
                        dfs.append(parse(df))
    if dfs:
        return pd.concat(dfs, ignore_index=True)
    else:
        return pd.DataFrame()


def create_excel(pdf_path):
    """
    #Convierte el DataFrame extraído del PDF en un Excel con formato.
    """
    try:
        df = read_table(pdf_path)
        if not df.empty:
            excel_path = os.path.splitext(pdf_path)[0] + ".xlsx"
            format_excel(df,"Movimientos","Movimientos",excel_path)
            return excel_path
    except PermissionError as e:
        print(f"Ocurrió un error\nProbablemente el archivo está abierto \n {e}")
    except Exception as e:
        print(f"Ocurrió un error \n{e}")




def create_all_excels(data):
    """
    Procesa la lista de dataFrames y y agrega la ruta de cada excel generado al diccionario data 
    """
    dat = []
    for temp in data:
        excel_path = create_excel(temp["pdf"])
        temp["excel"] = excel_path
        dat.append(temp)
    return dat



def format_excel(df, title, table_name, path):
    """
    Exporta un DataFrame a Excel y lo convierte en una tabla estructurada.
    Si el archivo existe, agrega una nueva hoja; si no, lo crea.
    Ajusta ancho de columnas y sobreescribe la tabla si ya existe.
    """

    # Si el archivo existe, lo cargamos; si no, lo creamos
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        # eliminar la hoja por defecto si no la quieres
        default_sheet = wb.active
        wb.remove(default_sheet)

    # Si la hoja ya existe, la eliminamos para reemplazarla
    if title in wb.sheetnames:
        ws_old = wb[title]
        wb.remove(ws_old)

    # Crear nueva hoja
    ws = wb.create_sheet(title)

    # Escribir encabezados
    ws.append(list(df.columns))

    # Escribir datos o fila dummy si está vacío
    if df.empty:
        # Agregar fila vacía para que Excel no elimine la tabla
        ws.append(["" for _ in df.columns])
        filas, columnas = 1, df.shape[1]
    else:
        for row in df.itertuples(index=False, name=None):
            ws.append(row)
        filas, columnas = df.shape

    # Definir rango de la tabla (incluye encabezado + filas)
    ref = f"A1:{openpyxl.utils.get_column_letter(columnas)}{filas+1}"

    # Si ya existe una tabla con ese nombre, eliminarla
    for t in ws._tables:
        if t.displayName == table_name:
            ws._tables.remove(t)

    # Crear tabla
    table = Table(displayName=table_name, ref=ref)

    # Estilo de tabla
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style

    # Agregar tabla a la hoja
    ws.add_table(table)

    # Aplicar formato en cursiva a las filas de datos
    italic_font = Font(italic=True)
    for row in ws.iter_rows(min_row=2, max_row=filas+1, min_col=1, max_col=columnas):
        for cell in row:
            cell.font = italic_font

    # Ajustar ancho de columnas
    for i, col in enumerate(df.columns, start=1):
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in ws[openpyxl.utils.get_column_letter(i)]
        )
        adjusted_width = max(max_length, len(col)) + 3
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = adjusted_width

    wb.save(path)
    return path
